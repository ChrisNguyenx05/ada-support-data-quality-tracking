from __future__ import annotations

import json
import math
import re
import tempfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import PatternFill


ROOT = Path(__file__).resolve().parents[1]
CONFIG_PATH = ROOT / "config" / "metric_mapping.json"
SELLER_COLORS = [
    "E8F3FF",
    "EAF8EE",
    "FFF4DE",
    "F3EAFE",
    "FFEAEA",
    "E9F7F8",
    "F8F0E7",
    "EEF2FF",
    "F1F5D8",
    "FCEBFA",
]


@dataclass
class CompareOptions:
    marketplace: str = "AUTO"
    seller_id: str = ""
    granularity: str = "day"
    platform_sheet: str | None = None


@dataclass
class SellerFileSpec:
    file_path: Path
    seller_id: str
    marketplace: str
    platform_sheet: str | None = None
    use_item_sales: bool = False


def load_config(path: Path = CONFIG_PATH) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def _clean_name(value: Any) -> str:
    text = "" if value is None else str(value)
    text = re.sub(r"\s+", " ", text.replace("\n", " ")).strip()
    return text


def _norm(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", _clean_name(value).lower())


def _first_matching_column(columns: list[str], aliases: list[str]) -> str | None:
    normalized = {_norm(col): col for col in columns}
    for alias in aliases:
        hit = normalized.get(_norm(alias))
        if hit is not None:
            return hit
    return None


def _read_csv(path: Path) -> pd.DataFrame:
    for encoding in ("utf-8-sig", "utf-8", "cp1252"):
        try:
            return pd.read_csv(path, encoding=encoding)
        except UnicodeDecodeError:
            continue
    return pd.read_csv(path)


def _find_header_row(raw: pd.DataFrame, aliases: dict[str, list[str]]) -> int:
    best_row = 0
    best_score = -1
    required = aliases.get("date", [])
    for idx in range(min(len(raw), 80)):
        row_values = [_clean_name(v) for v in raw.iloc[idx].tolist()]
        row_norm = {_norm(v) for v in row_values if v}
        score = 0
        for metric, names in aliases.items():
            if any(_norm(name) in row_norm for name in names):
                score += 3 if metric in ("date", "seller_id") else 1
        if required and not any(_norm(name) in row_norm for name in required):
            score -= 2
        if score > best_score:
            best_score = score
            best_row = idx
    return best_row


def _read_excel_tables(path: Path, aliases: dict[str, list[str]], sheet: str | None = None) -> list[tuple[str, pd.DataFrame]]:
    if path.suffix.lower() == ".xls":
        raise ValueError("File .xls canh cu khong doc truc tiep duoc trong tool nay. Hay Save As .xlsx hoac export CSV.")

    xl = pd.ExcelFile(path)
    sheet_names = [sheet] if sheet else xl.sheet_names
    tables: list[tuple[str, pd.DataFrame]] = []
    for sheet_name in sheet_names:
        raw = pd.read_excel(path, sheet_name=sheet_name, header=None, dtype=object)
        if raw.empty:
            continue
        header_idx = _find_header_row(raw, aliases)
        headers = [_clean_name(v) or f"Column {i + 1}" for i, v in enumerate(raw.iloc[header_idx].tolist())]
        table = raw.iloc[header_idx + 1 :].copy()
        table.columns = headers
        table = table.dropna(how="all")
        table = table.loc[:, [col for col in table.columns if not str(col).startswith("Column ")]]
        if not table.empty:
            tables.append((sheet_name, table))
    return tables


def read_any_table(path: Path, aliases: dict[str, list[str]], sheet: str | None = None) -> list[tuple[str, pd.DataFrame]]:
    suffix = path.suffix.lower()
    if suffix in (".csv", ".tsv"):
        df = _read_csv(path) if suffix == ".csv" else pd.read_csv(path, sep="\t")
        return [(path.stem, df)]
    if suffix in (".xlsx", ".xlsm", ".xls"):
        return _read_excel_tables(path, aliases, sheet)
    raise ValueError(f"Khong ho tro file {suffix}. Hay dung .xlsx, .csv, hoac .tsv.")


def _parse_date(value: Any) -> pd.Timestamp | pd.NaT:
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return pd.NaT
    text = _clean_name(value)
    if not text:
        return pd.NaT
    date_hits = re.findall(r"\d{1,4}[./-]\d{1,2}[./-]\d{1,4}", text)
    if len(date_hits) >= 2:
        text = date_hits[0]
    dayfirst_order = (True, False) if re.match(r"^\d{1,2}/\d{1,2}/\d{4}$", text) else (False, True)
    for dayfirst in dayfirst_order:
        parsed = pd.to_datetime(text, errors="coerce", dayfirst=dayfirst)
        if pd.notna(parsed):
            return parsed.normalize()
    return pd.NaT


def _period(series: pd.Series, granularity: str) -> pd.Series:
    dates = series.map(_parse_date)
    if granularity == "month":
        return dates.dt.to_period("M").astype(str)
    if granularity == "week":
        return dates.dt.to_period("W-MON").astype(str)
    return dates.dt.strftime("%Y-%m-%d")


def _is_date_range(value: Any) -> bool:
    text = _clean_name(value)
    return len(re.findall(r"\d{1,4}[./-]\d{1,2}[./-]\d{1,4}", text)) >= 2


def _to_number(value: Any) -> float:
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = _clean_name(value)
    if not text:
        return 0.0
    text = text.replace(",", "").replace("%", "")
    text = re.sub(r"[^\d.\-]", "", text)
    if text in ("", "-", ".", "-."):
        return 0.0
    return float(text)


def _normalize_table(
    df: pd.DataFrame,
    aliases: dict[str, list[str]],
    canonical_metrics: dict[str, Any],
    granularity: str,
    seller_id: str = "",
) -> tuple[pd.DataFrame, dict[str, str]]:
    columns = [_clean_name(c) for c in df.columns]
    df = df.copy()
    df.columns = columns

    mapping: dict[str, str] = {}
    date_col = _first_matching_column(columns, aliases.get("date", []))
    if date_col is None:
        raise ValueError("Khong tim thay cot Date/Day trong file.")
    mapping["date"] = date_col

    seller_col = _first_matching_column(columns, aliases.get("seller_id", []))
    if seller_col:
        mapping["seller_id"] = seller_col

    range_mask = df[date_col].map(_is_date_range)
    if granularity == "month" and range_mask.any():
        df = df[range_mask].copy()
    elif granularity == "day" and range_mask.any():
        df = df[~range_mask].copy()

    region_col = _first_matching_column(columns, ["Region"])
    if region_col and df[region_col].map(_clean_name).str.upper().eq("ALL").any():
        non_all = ~df[region_col].map(_clean_name).str.upper().eq("ALL")
        if non_all.any():
            df = df[non_all].copy()

    normalized = pd.DataFrame()
    normalized["period"] = _period(df[date_col], granularity)
    normalized["seller_id"] = df[seller_col].map(_clean_name) if seller_col else seller_id.strip()

    for metric in canonical_metrics:
        col = _first_matching_column(columns, aliases.get(metric, []))
        if col:
            mapping[metric] = col
            normalized[metric] = df[col].map(_to_number)

    normalized = normalized[normalized["period"].notna() & (normalized["period"] != "NaT")]
    metric_cols = [c for c in canonical_metrics if c in normalized.columns]
    if not metric_cols:
        raise ValueError("Khong tim thay metric nao de so sanh. Hay cap nhat config/metric_mapping.json.")
    if seller_id.strip():
        normalized["seller_id"] = seller_id.strip()
    normalized = normalized[["seller_id", "period", *metric_cols]]
    return normalized, mapping


def _aggregate(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    metric_cols = [c for c in df.columns if c not in ("seller_id", "period")]
    key_cols = ["seller_id", "period"]
    duplicates = (
        df.groupby(key_cols, dropna=False)
        .size()
        .reset_index(name="row_count")
        .query("row_count > 1")
    )
    aggregated = df.groupby(key_cols, dropna=False)[metric_cols].sum().reset_index()
    return aggregated, duplicates


def read_platform_file(spec: SellerFileSpec, granularity: str) -> tuple[pd.DataFrame, list[dict[str, str]], list[str]]:
    config = load_config()
    marketplace = (spec.marketplace or "AUTO").upper()
    platform_aliases = config["platform_aliases"].get(marketplace, config["platform_aliases"]["AUTO"])
    canonical = config["canonical_metrics"]
    tables = read_any_table(spec.file_path, platform_aliases, spec.platform_sheet)
    errors: list[str] = []
    parts: list[pd.DataFrame] = []
    mappings: list[dict[str, str]] = []
    for sheet_name, table in tables:
        try:
            norm, mapping = _normalize_table(table, platform_aliases, canonical, granularity, spec.seller_id)
            norm.insert(0, "file_name", spec.file_path.name)
            norm.insert(1, "marketplace", marketplace)
            norm.insert(2, "source_sheet", sheet_name)
            parts.append(norm)
            mappings.append({"file_name": spec.file_path.name, "seller_id": spec.seller_id, "marketplace": marketplace, "sheet": sheet_name, **mapping})
        except Exception as exc:
            errors.append(f"{spec.file_path.name}/{sheet_name}: {exc}")
    if not parts:
        raise ValueError("Khong doc duoc platform file. " + " | ".join(errors))
    return pd.concat(parts, ignore_index=True), mappings, errors


def _to_period_from_day(series: pd.Series, granularity: str) -> pd.Series:
    dates = pd.to_datetime(series, errors="coerce")
    if granularity == "month":
        return dates.dt.to_period("M").astype(str)
    if granularity == "week":
        return dates.dt.to_period("W-MON").astype(str)
    return dates.dt.strftime("%Y-%m-%d")


def _is_close(left: float, right: float, metric: str, canonical: dict[str, Any]) -> bool:
    tolerance_abs = canonical[metric].get("tolerance_abs", 0)
    tolerance_pct = canonical[metric].get("tolerance_pct", 0)
    diff = right - left
    pct = diff / left if left else (0 if right == 0 else 1)
    return abs(diff) <= tolerance_abs or abs(pct) <= tolerance_pct


def _safe_sheet_name(value: Any, prefix: str = "") -> str:
    text = _clean_name(value) or "blank"
    text = re.sub(r"[\[\]:*?/\\]", "_", text)
    return f"{prefix}{text}"[:31]


def _style_workbook_by_seller(writer: pd.ExcelWriter, sheet_names: list[str]) -> None:
    seller_colors: dict[str, str] = {}
    for sheet_name in sheet_names:
        worksheet = writer.sheets.get(sheet_name)
        if worksheet is None:
            continue
        headers = [cell.value for cell in worksheet[1]]
        if "seller_id" not in headers:
            continue
        seller_col = headers.index("seller_id") + 1
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        for row in range(2, worksheet.max_row + 1):
            seller_id = worksheet.cell(row=row, column=seller_col).value
            if not seller_id:
                continue
            seller_id = str(seller_id)
            if seller_id not in seller_colors:
                seller_colors[seller_id] = SELLER_COLORS[len(seller_colors) % len(SELLER_COLORS)]
            fill = PatternFill("solid", fgColor=seller_colors[seller_id])
            for col in range(1, worksheet.max_column + 1):
                worksheet.cell(row=row, column=col).fill = fill


def compare_platform_to_db_sources(
    platform_norm: pd.DataFrame,
    db_by_source: pd.DataFrame,
    granularity: str,
    output_dir: Path | None = None,
    data_level: str = "both",
) -> dict[str, Any]:
    config = load_config()
    canonical = config["canonical_metrics"]
    metrics = ["quantity", "revenue", "page_view"]

    platform_cols = ["seller_id", "period", *[m for m in metrics if m in platform_norm.columns]]
    platform_agg = platform_norm[platform_cols].groupby(["seller_id", "period"], dropna=False).sum(numeric_only=True).reset_index()

    db = db_by_source.copy()
    if db.empty:
        db = pd.DataFrame(columns=["data_type", "data_level", "seller_id", "day", "source", *metrics])
    db["period"] = _to_period_from_day(db["day"], granularity) if "day" in db.columns else pd.Series(dtype=str)
    for metric in metrics:
        if metric not in db.columns:
            db[metric] = 0.0
        db[metric] = pd.to_numeric(db[metric], errors="coerce").fillna(0.0)

    db_all = (
        db.groupby(["seller_id", "period", "data_level", "data_type"], dropna=False)[metrics]
        .sum()
        .reset_index()
    )
    db_source = (
        db.groupby(["seller_id", "period", "data_level", "data_type", "source"], dropna=False)[metrics]
        .sum()
        .reset_index()
    )

    detail_rows: list[dict[str, Any]] = []
    source_rows: list[dict[str, Any]] = []
    target_index = platform_agg.set_index(["seller_id", "period"])
    expected_rows = [
        {"data_level": "seller", "data_type": "seller_sales"},
        {"data_level": "seller", "data_type": "seller_traffic"},
        {"data_level": "sku", "data_type": "sku_sales"},
        {"data_level": "sku", "data_type": "sku_traffic"},
    ]
    wanted_level = (data_level or "both").lower()
    if wanted_level in ("seller", "sku"):
        expected_rows = [row for row in expected_rows if row["data_level"] == wanted_level]
        db = db[db["data_level"] == wanted_level].copy() if not db.empty else db
        db_all = db_all[db_all["data_level"] == wanted_level].copy() if not db_all.empty else db_all
        db_source = db_source[db_source["data_level"] == wanted_level].copy() if not db_source.empty else db_source
    expected_types = pd.DataFrame(expected_rows)
    platform_keys = platform_agg[["seller_id", "period"]].drop_duplicates()
    db_keys = platform_keys.merge(expected_types, how="cross")

    for _, key in db_keys.iterrows():
        seller_id = key["seller_id"]
        period = key["period"]
        data_level = key["data_level"]
        data_type = key["data_type"]
        if (seller_id, period) not in target_index.index:
            continue
        target = target_index.loc[(seller_id, period)]
        if isinstance(target, pd.DataFrame):
            target = target.iloc[0]
        all_slice = db_all[
            (db_all["seller_id"] == seller_id)
            & (db_all["period"] == period)
            & (db_all["data_level"] == data_level)
            & (db_all["data_type"] == data_type)
        ]
        all_row = all_slice.iloc[0] if not all_slice.empty else pd.Series({"quantity": 0.0, "revenue": 0.0, "page_view": 0.0})
        source_slice = db_source[
            (db_source["seller_id"] == seller_id)
            & (db_source["period"] == period)
            & (db_source["data_level"] == data_level)
            & (db_source["data_type"] == data_type)
        ]
        metric_list = [m for m in metrics if m in platform_agg.columns and not data_type.endswith("traffic") or m == "page_view"]
        if data_type.endswith("sales"):
            metric_list = [m for m in ("quantity", "revenue") if m in platform_agg.columns]
        elif data_type.endswith("traffic"):
            metric_list = [m for m in ("page_view",) if m in platform_agg.columns]

        for metric in metric_list:
            platform_value = float(target.get(metric, 0) or 0)
            db_value = float(all_row.get(metric, 0) or 0)
            all_match = _is_close(platform_value, db_value, metric, canonical)
            matching_sources = []
            for _, source_row in source_slice.iterrows():
                source_value = float(source_row.get(metric, 0) or 0)
                source_match = _is_close(platform_value, source_value, metric, canonical)
                if source_match:
                    matching_sources.append(str(source_row["source"]))
                source_diff = source_value - platform_value
                source_rows.append(
                    {
                        "seller_id": seller_id,
                        "period": period,
                        "data_level": data_level,
                        "data_type": data_type,
                        "metric": metric,
                        "source": source_row["source"],
                        "source_value": source_value,
                        "platform_value": platform_value,
                        "source_diff": source_diff,
                        "source_accuracy": source_value / platform_value if platform_value else (1 if source_value == 0 else 0),
                        "source_match": source_match,
                    }
                )
            if all_match:
                status = "match"
                recommendation = "OK"
            elif matching_sources:
                status = "suspicious_extra_source"
                recommendation = "Mot source da khop platform nhung tong all sources bi du. Can Tech review/remove source con lai."
            elif db_value == 0 and platform_value != 0:
                status = "missing_in_db"
                recommendation = "DB khong co data cho metric nay. Can rerun/check transform."
            else:
                status = "mismatch"
                recommendation = "DB co data nhung khong source nao khop platform. Can check collect/source logic."
            diff = db_value - platform_value
            detail_rows.append(
                {
                    "seller_id": seller_id,
                    "period": period,
                    "data_level": data_level,
                    "data_type": data_type,
                    "metric": metric,
                    "platform_value": platform_value,
                    "db_all_sources": db_value,
                    "diff_db_minus_platform": diff,
                    "diff_pct": diff / platform_value if platform_value else (0 if db_value == 0 else 1),
                    "status": status,
                    "matching_source_alone": ", ".join(matching_sources),
                    "checked_source": True,
                    "recommendation": recommendation,
                }
            )

    detail = pd.DataFrame(detail_rows)
    source_detail = pd.DataFrame(source_rows)
    if not source_detail.empty:
        source_detail["investigation_note"] = source_detail.apply(
            lambda row: "MATCH_UI" if row["source_match"] else ("OVER_UI" if row["source_value"] > row["platform_value"] else "UNDER_UI"),
            axis=1,
        )
    duplicate_source = (
        db.groupby(["seller_id", "period", "data_level", "data_type"], dropna=False)["source"]
        .nunique()
        .reset_index(name="source_count")
        .query("source_count > 1")
    )
    summary = detail.groupby("status").size().reset_index(name="count") if not detail.empty else pd.DataFrame(columns=["status", "count"])

    output_dir = output_dir or ROOT / "outputs"
    output_dir.mkdir(parents=True, exist_ok=True)
    out_path = output_dir / f"dq_batch_report_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Summary", index=False)
        wrong_detail = detail[detail["status"] != "match"] if "status" in detail.columns else detail
        wrong_detail.to_excel(writer, sheet_name="Wrong_Data_Detail", index=False)
        source_detail.to_excel(writer, sheet_name="Source_Investigation", index=False)
        source_detail.to_excel(writer, sheet_name="Source_Check_Detail", index=False)
        duplicate_source.to_excel(writer, sheet_name="Duplicate_Source", index=False)
        (detail[detail["status"] == "missing_in_db"] if "status" in detail.columns else detail).to_excel(writer, sheet_name="Missing_In_DB", index=False)
        (detail[detail["status"] == "mismatch"] if "status" in detail.columns else detail).to_excel(writer, sheet_name="Mismatch_Value", index=False)
        (detail[detail["status"] == "match"] if "status" in detail.columns else detail).to_excel(writer, sheet_name="Matched", index=False)
        platform_norm.to_excel(writer, sheet_name="Raw_Platform_Normalized", index=False)
        db.to_excel(writer, sheet_name="Raw_DB_By_Source", index=False)
        per_seller_sheet_names = []
        if not wrong_detail.empty and "seller_id" in wrong_detail.columns:
            for seller_id, seller_rows in list(wrong_detail.groupby("seller_id", dropna=False))[:20]:
                sheet_name = _safe_sheet_name(seller_id, "S_")
                seller_rows.to_excel(writer, sheet_name=sheet_name, index=False)
                per_seller_sheet_names.append(sheet_name)
        _style_workbook_by_seller(
            writer,
            [
                "Wrong_Data_Detail",
                "Source_Investigation",
                "Source_Check_Detail",
                "Duplicate_Source",
                "Missing_In_DB",
                "Mismatch_Value",
                "Matched",
                "Raw_Platform_Normalized",
                "Raw_DB_By_Source",
                *per_seller_sheet_names,
            ],
        )

    return {
        "summary": summary.to_dict(orient="records"),
        "comparison": detail.sort_values(["status", "seller_id", "period", "data_level", "data_type", "metric"]).head(500).to_dict(orient="records") if not detail.empty else [],
        "report_path": str(out_path.resolve()),
        "source_rows": len(source_detail),
        "duplicate_source": len(duplicate_source),
    }


def compare_files(
    platform_path: Path,
    db_path: Path,
    output_dir: Path | None = None,
    options: CompareOptions | None = None,
) -> dict[str, Any]:
    options = options or CompareOptions()
    config = load_config()
    marketplace = (options.marketplace or "AUTO").upper()
    platform_aliases = config["platform_aliases"].get(marketplace, config["platform_aliases"]["AUTO"])
    db_aliases = config["db_aliases"]
    canonical = config["canonical_metrics"]

    platform_tables = read_any_table(platform_path, platform_aliases, options.platform_sheet)
    platform_errors: list[str] = []
    platform_norm_parts: list[pd.DataFrame] = []
    platform_mappings: list[dict[str, str]] = []
    for sheet_name, table in platform_tables:
        try:
            norm, mapping = _normalize_table(table, platform_aliases, canonical, options.granularity, options.seller_id)
            norm.insert(0, "source_sheet", sheet_name)
            platform_norm_parts.append(norm)
            platform_mappings.append({"sheet": sheet_name, **mapping})
        except Exception as exc:
            platform_errors.append(f"{sheet_name}: {exc}")
    if not platform_norm_parts:
        raise ValueError("Khong doc duoc bang platform nao. " + " | ".join(platform_errors))
    platform_norm = pd.concat(platform_norm_parts, ignore_index=True)

    db_tables = read_any_table(db_path, db_aliases)
    db_norm_parts: list[pd.DataFrame] = []
    db_errors: list[str] = []
    for sheet_name, table in db_tables:
        try:
            norm, _ = _normalize_table(table, db_aliases, canonical, options.granularity, options.seller_id)
            norm.insert(0, "source_sheet", sheet_name)
            db_norm_parts.append(norm)
        except Exception as exc:
            db_errors.append(f"{sheet_name}: {exc}")
    if not db_norm_parts:
        raise ValueError("Khong doc duoc bang DB export nao. " + " | ".join(db_errors))
    db_norm = pd.concat(db_norm_parts, ignore_index=True)

    platform_for_agg = platform_norm.drop(columns=["source_sheet"], errors="ignore")
    db_for_agg = db_norm.drop(columns=["source_sheet"], errors="ignore")
    platform_agg, platform_dupes = _aggregate(platform_for_agg)
    db_agg, db_dupes = _aggregate(db_for_agg)

    metrics = sorted((set(platform_agg.columns) & set(db_agg.columns)) - {"seller_id", "period"})
    merged = platform_agg.merge(db_agg, on=["seller_id", "period"], how="outer", suffixes=("_platform", "_db"), indicator=True)
    rows: list[dict[str, Any]] = []
    for _, row in merged.iterrows():
        for metric in metrics:
            pv_raw = row.get(f"{metric}_platform", 0)
            dv_raw = row.get(f"{metric}_db", 0)
            pv = 0.0 if pd.isna(pv_raw) else float(pv_raw)
            dv = 0.0 if pd.isna(dv_raw) else float(dv_raw)
            diff = dv - pv
            pct = diff / pv if pv else (0 if dv == 0 else 1)
            tolerance_abs = canonical[metric].get("tolerance_abs", 0)
            tolerance_pct = canonical[metric].get("tolerance_pct", 0)
            if row["_merge"] == "left_only":
                status = "missing_in_db"
            elif row["_merge"] == "right_only":
                status = "missing_in_platform"
            elif abs(diff) <= tolerance_abs or abs(pct) <= tolerance_pct:
                status = "match"
            else:
                status = "mismatch"
            rows.append(
                {
                    "seller_id": row["seller_id"],
                    "period": row["period"],
                    "metric": metric,
                    "platform_value": pv,
                    "db_value": dv,
                    "diff_db_minus_platform": diff,
                    "diff_pct": pct,
                    "status": status,
                }
            )
    comparison = pd.DataFrame(rows)
    summary = comparison.groupby("status").size().reset_index(name="count") if not comparison.empty else pd.DataFrame()

    output_dir = output_dir or ROOT / "outputs"
    output_dir.mkdir(parents=True, exist_ok=True)
    out_path = output_dir / f"dq_report_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Summary", index=False)
        comparison.to_excel(writer, sheet_name="Comparison", index=False)
        platform_dupes.to_excel(writer, sheet_name="Platform_Duplicates", index=False)
        db_dupes.to_excel(writer, sheet_name="DB_Duplicates", index=False)
        platform_norm.to_excel(writer, sheet_name="Platform_Normalized", index=False)
        db_norm.to_excel(writer, sheet_name="DB_Normalized", index=False)
        pd.DataFrame(platform_mappings).to_excel(writer, sheet_name="Column_Mapping", index=False)

    return {
        "summary": summary.to_dict(orient="records"),
        "comparison": comparison.sort_values(["status", "period", "metric"]).head(300).to_dict(orient="records"),
        "report_path": str(out_path.resolve()),
        "platform_errors": platform_errors,
        "db_errors": db_errors,
        "platform_duplicates": len(platform_dupes),
        "db_duplicates": len(db_dupes),
        "metrics": metrics,
    }


def save_upload(field: Any, upload_dir: Path) -> Path:
    filename = Path(field.filename or "upload").name
    target = upload_dir / filename
    target.write_bytes(field.file.read())
    return target


def temp_upload_dir() -> Path:
    return Path(tempfile.mkdtemp(prefix="dq_checker_"))
